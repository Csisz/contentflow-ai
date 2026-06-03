import json

from openpyxl import Workbook

from contentflow_ai.migration.config import load_config
from contentflow_ai.migration.excel_parser import normalize_category_value, parse_workbook


def _write_config(path, local_root=""):
    path.write_text(json.dumps({
        "base_url": "https://example/otcs/cs.exe",
        "username": "technical_user",
        "password": "secret",
        "enterprise_node_id": 2000,
        "category_id": 123,
        "ws_sheet": "Workspace",
        "file_sheet": "File",
        "ws_data_start_row": 1,
        "file_data_start_row": 1,
        "local_file_root": local_root,
        "ws_columns": {"location": 0, "title": 1},
        "file_columns": {"location": 0, "title": 1, "src": 2, "mime": 3},
        "category_fields": {
            "doctype": {"attr_id": 3, "col": 2, "required": True, "value_map": {"Spec": "Specification"}},
            "country": {"attr_id": 22, "multi_value": True, "col_start": 3, "col_end": 4}
        }
    }), encoding="utf-8")


def test_parse_workbook_reads_workspace_and_file_rows(tmp_path):
    source = tmp_path / "source.pdf"
    source.write_bytes(b"fake")
    wb = Workbook()
    ws = wb.active
    ws.title = "Workspace"
    ws.append(["location", "title", "doctype", "country1", "country2"])
    ws.append(["Enterprise/Test", "WS-001", "Spec", "HU", "DE"])
    files = wb.create_sheet("File")
    files.append(["location", "title", "src", "mime"])
    files.append(["Enterprise/Test/WS-001", "doc.pdf", str(source), "pdf"])
    xlsx = tmp_path / "migration.xlsx"
    wb.save(xlsx)

    config_path = tmp_path / "config.json"
    _write_config(config_path)
    cfg = load_config(config_path)

    parsed = parse_workbook(xlsx, cfg)

    assert len(parsed.workspaces) == 1
    assert parsed.workspaces[0].cat_values["doctype"] == "Specification"
    assert parsed.workspaces[0].cat_values["country"] == ["HU", "DE"]
    assert len(parsed.files) == 1
    assert parsed.files[0].local_path == str(source)


def test_category_value_map_normalizes_single_value_variants(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Workspace"
    ws.append(["location", "title", "pkg_type"])
    variants = [
        "Címke",
        "Ci\u0301mke",
        " Címke ",
        "Címke\u00a0",
        "címke",
        "CÍMKE",
    ]
    for index, value in enumerate(variants, start=1):
        ws.append(["Enterprise/Test", f"WS-{index:03d}", value])
    files = wb.create_sheet("File")
    files.append(["location", "title", "src"])
    xlsx = tmp_path / "migration.xlsx"
    wb.save(xlsx)

    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps({
        "base_url": "https://example/otcs/cs.exe",
        "username": "technical_user",
        "password": "secret",
        "enterprise_node_id": 2000,
        "category_id": 123,
        "ws_sheet": "Workspace",
        "file_sheet": "File",
        "ws_data_start_row": 1,
        "file_data_start_row": 1,
        "ws_columns": {"location": 0, "title": 1},
        "file_columns": {"location": 0, "title": 1, "src": 2},
        "category_fields": {
            "pkg_type": {
                "attr_id": 3,
                "col": 2,
                "value_map": {"Címke": "Címke / Label"},
            }
        },
    }), encoding="utf-8")
    cfg = load_config(config_path)

    parsed = parse_workbook(xlsx, cfg)

    assert [ws.cat_values["pkg_type"] for ws in parsed.workspaces] == ["Címke / Label"] * len(variants)


def test_category_value_map_normalizes_multi_value_items(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Workspace"
    ws.append(["location", "title", "pkg_type_1", "pkg_type_2"])
    ws.append(["Enterprise/Test", "WS-001", " Címke ", "doboz"])
    files = wb.create_sheet("File")
    files.append(["location", "title", "src"])
    xlsx = tmp_path / "migration.xlsx"
    wb.save(xlsx)

    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps({
        "base_url": "https://example/otcs/cs.exe",
        "username": "technical_user",
        "password": "secret",
        "enterprise_node_id": 2000,
        "category_id": 123,
        "ws_sheet": "Workspace",
        "file_sheet": "File",
        "ws_data_start_row": 1,
        "file_data_start_row": 1,
        "ws_columns": {"location": 0, "title": 1},
        "file_columns": {"location": 0, "title": 1, "src": 2},
        "category_fields": {
            "pkg_type": {
                "attr_id": 3,
                "multi_value": True,
                "col_start": 2,
                "col_end": 3,
                "value_map": {
                    "Címke": "Címke / Label",
                    "Doboz": "Doboz / Box",
                },
            }
        },
    }), encoding="utf-8")
    cfg = load_config(config_path)

    parsed = parse_workbook(xlsx, cfg)

    assert parsed.workspaces[0].cat_values["pkg_type"] == ["Címke / Label", "Doboz / Box"]


def test_normalize_category_value_collapses_whitespace_and_casefolds():
    assert normalize_category_value(" Ci\u0301mke\u00a0  extra ") == "címke extra"
