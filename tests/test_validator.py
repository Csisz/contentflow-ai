import json
from pathlib import Path

from openpyxl import Workbook

from contentflow_ai.migration.config import load_config
from contentflow_ai.migration.excel_parser import parse_workbook
from contentflow_ai.migration.cs_client import NodeInfo
from contentflow_ai.migration.models import CategoryFieldConfig, MigrationWorkbook, WorkspaceRow
from contentflow_ai.migration.validator import PreflightValidator


def test_validator_flags_missing_file_and_duplicate_target(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Workspace"
    ws.append(["location", "title", "doctype"])
    ws.append(["Enterprise/Test", "WS-001", "Contract"])
    files = wb.create_sheet("File")
    files.append(["location", "title", "src", "mime"])
    files.append(["Enterprise/Test/WS-001", "doc.pdf", "missing.pdf", "pdf"])
    files.append(["Enterprise/Test/WS-001", "doc.pdf", "missing.pdf", "pdf"])
    xlsx = tmp_path / "migration.xlsx"
    wb.save(xlsx)

    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps({
        "base_url": "https://example/otcs/cs.exe",
        "username": "technical_user",
        "password": "secret",
        "enterprise_node_id": 2000,
        "category_id": 123,
        "ws_sheet": "Workspace",
        "file_sheet": "File",
        "ws_data_start_row": 1,
        "file_data_start_row": 1,
        "ws_columns": {"location": 0, "title": 1},
        "file_columns": {"location": 0, "title": 1, "src": 2, "mime": 3},
        "category_fields": {"doctype": {"attr_id": 3, "col": 2, "required": True}}
    }), encoding="utf-8")
    cfg = load_config(config_path)
    workbook = parse_workbook(xlsx, cfg)

    report = PreflightValidator(cfg).validate(workbook)
    codes = {issue.code for issue in report.issues}

    assert "MISSING_SOURCE_FILE" in codes
    assert "DUPLICATE_FILE_IN_TARGET" in codes
    assert report.summary.decision == "NO_GO"
    assert report.summary.readiness_score < 100


class PlanningFakeCSClient:
    def __init__(self, root_id: int):
        self.root_id = root_id

    def authenticate(self) -> None:
        return None

    def get_node(self, node_id: int) -> NodeInfo | None:
        if node_id == self.root_id:
            return NodeInfo(node_id, "MIGR", 0)
        return None

    def plan_path_creation(self, location: str) -> dict:
        return {
            "requested_path": location,
            "relative_path": "Client/New",
            "root_node_id": self.root_id,
            "root_name": "MIGR",
            "existing_until_node_id": self.root_id,
            "existing_until_path": "MIGR",
            "missing_parts": ["Client", "New"],
            "full_path_exists": False,
            "action": "create_missing_folders",
        }


def test_preflight_missing_target_is_planned_creation_info(tmp_path):
    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps({
        "base_url": "https://example/otcs/cs.exe",
        "username": "technical_user",
        "password": "secret",
        "enterprise_node_id": 2000,
        "category_id": 123,
        "ws_sheet": "Workspace",
        "file_sheet": "File",
        "ws_columns": {"location": 0, "title": 1},
        "file_columns": {"location": 0, "title": 1, "src": 2},
        "category_fields": {}
    }), encoding="utf-8")
    cfg = load_config(config_path)
    workbook = MigrationWorkbook(
        xlsx_path=Path("migration.xlsx"),
        workspaces=[WorkspaceRow(row_index=2, location="Enterprise/MIGR/Client/New", title="WS-001")],
        files=[],
        sheet_names=["Workspace", "File"],
    )

    report = PreflightValidator(cfg, cs_client=PlanningFakeCSClient(2000)).validate(
        workbook,
        include_content_server=True,
    )
    issues_by_code = {issue.code: issue for issue in report.issues}

    assert "TARGET_LOCATION_MISSING" not in issues_by_code
    assert issues_by_code["TARGET_WILL_BE_CREATED"].severity == "info"
    assert issues_by_code["TARGET_WILL_BE_CREATED"].row_type == "content_server"
    assert issues_by_code["TARGET_WILL_BE_CREATED"].field == "location"
    assert issues_by_code["TARGET_WILL_BE_CREATED"].value == "Enterprise/MIGR/Client/New"
    assert "MIGR (2000)" in issues_by_code["TARGET_WILL_BE_CREATED"].suggestion
    assert "Client, New" in issues_by_code["TARGET_WILL_BE_CREATED"].suggestion


def test_preflight_warns_when_category_value_map_does_not_map_value():
    cfg = load_config_from_dict({
        "base_url": "https://example/otcs/cs.exe",
        "username": "technical_user",
        "password": "secret",
        "enterprise_node_id": 2000,
        "category_id": 123,
        "ws_sheet": "Workspace",
        "file_sheet": "File",
        "ws_columns": {"location": 0, "title": 1},
        "file_columns": {"location": 0, "title": 1, "src": 2},
        "category_fields": {},
    })
    cfg.category_fields = {
        "pkg_type": CategoryFieldConfig(
            key="pkg_type",
            attr_id=3,
            value_map={"Címke": "Címke / Label"},
        )
    }
    workbook = MigrationWorkbook(
        xlsx_path=Path("migration.xlsx"),
        workspaces=[
            WorkspaceRow(
                row_index=7,
                location="Enterprise/Test",
                title="WS-001",
                cat_values={"pkg_type": "Tasak"},
            ),
            WorkspaceRow(
                row_index=8,
                location="Enterprise/Test",
                title="WS-002",
                cat_values={"pkg_type": "Címke / Label"},
            ),
        ],
        files=[],
        sheet_names=["Workspace", "File"],
    )

    report = PreflightValidator(cfg).validate(workbook)
    warnings = [issue for issue in report.issues if issue.code == "CATEGORY_VALUE_NOT_MAPPED"]

    assert len(warnings) == 1
    assert warnings[0].severity == "warning"
    assert warnings[0].row_type == "workspace"
    assert warnings[0].field == "pkg_type"
    assert warnings[0].message == "Category value has no value_map entry and will be sent as-is."
    assert warnings[0].value == "Tasak"


def load_config_from_dict(data, tmp_path=None):
    import tempfile

    path = Path(tempfile.mkdtemp()) / "config.json"
    path.write_text(json.dumps(data), encoding="utf-8")
    return load_config(path)
