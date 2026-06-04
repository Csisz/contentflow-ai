from __future__ import annotations

import os
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import MigrationConfig
from .models import FileRow, MigrationWorkbook, RelatedWorkspaceRow, WorkspaceRow
from .utils import clean_cell_value

_WHITESPACE_PATTERN = re.compile(r"\s+")
RELATED_WORKSPACE_SHEET = "RelatedWorkspace"
ENABLED_TRUE_VALUES = {"1", "true", "yes", "y", "igen"}


class WorkbookParseError(ValueError):
    pass


def parse_workbook(xlsx_path: str | Path, cfg: MigrationConfig) -> MigrationWorkbook:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"XLSX not found: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    missing = [name for name in (cfg.ws_sheet, cfg.file_sheet) if name not in sheet_names]
    if missing:
        raise WorkbookParseError(f"Missing required sheet(s): {', '.join(missing)}")

    workspaces = _parse_workspaces(wb[cfg.ws_sheet], cfg)
    files = _parse_files(wb[cfg.file_sheet], cfg)
    related_workspaces = _parse_related_workspaces(wb[RELATED_WORKSPACE_SHEET]) if RELATED_WORKSPACE_SHEET in sheet_names else []
    return MigrationWorkbook(path, workspaces, files, sheet_names, related_workspaces)


def _parse_workspaces(sheet: Any, cfg: MigrationConfig) -> list[WorkspaceRow]:
    rows: list[WorkspaceRow] = []
    location_col = cfg.ws_columns.get("location")
    title_col = cfg.ws_columns.get("title")
    for excel_row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        zero_based_index = excel_row_number - 1
        if zero_based_index < cfg.ws_data_start_row:
            continue
        location = _safe(row, location_col)
        title = _safe(row, title_col)
        if not title and not location:
            continue
        cat_values: dict[str, Any] = {}
        for key, field_cfg in cfg.category_fields.items():
            if field_cfg.multi_value:
                start = field_cfg.col_start if field_cfg.col_start is not None else field_cfg.col
                end = field_cfg.col_end if field_cfg.col_end is not None else start
                values = [_safe(row, col) for col in range(int(start), int(end) + 1)] if start is not None else []
                if field_cfg.value_map:
                    values = [_apply_value_map(value, field_cfg.value_map) for value in values]
                cat_values[key] = [value for value in values if value]
            else:
                value = _safe(row, field_cfg.col)
                if value and field_cfg.value_map:
                    value = _apply_value_map(value, field_cfg.value_map)
                cat_values[key] = value
        rows.append(WorkspaceRow(row_index=excel_row_number, location=location, title=title, cat_values=cat_values))
    return rows


def _parse_files(sheet: Any, cfg: MigrationConfig) -> list[FileRow]:
    rows: list[FileRow] = []
    location_col = cfg.file_columns.get("location")
    title_col = cfg.file_columns.get("title")
    src_col = cfg.file_columns.get("src", cfg.file_columns.get("file"))
    mime_col = cfg.file_columns.get("mime")
    version_col = cfg.file_columns.get("version")
    skip_locations = {"location", "file adatok", "title", "file név", ""}

    for excel_row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        zero_based_index = excel_row_number - 1
        if zero_based_index < cfg.file_data_start_row:
            continue
        title = _safe(row, title_col)
        location = _safe(row, location_col)
        src = _safe(row, src_col)
        if not title and not location and not src:
            continue
        if location.lower() in skip_locations and not src:
            continue
        local_path = _resolve_local_path(src, cfg.local_file_root)
        rows.append(
            FileRow(
                row_index=excel_row_number,
                location=location,
                title=title,
                src=src,
                local_path=local_path,
                mime_hint=_safe(row, mime_col),
                version=_safe(row, version_col),
            )
        )
    return rows


def _parse_related_workspaces(sheet: Any) -> list[RelatedWorkspaceRow]:
    rows: list[RelatedWorkspaceRow] = []
    header_map: dict[str, int] | None = None

    for excel_row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if header_map is None:
            header_map = {_safe_header(value): index for index, value in enumerate(row) if _safe_header(value)}
            continue

        source_workspace = _safe(row, header_map.get("source_workspace"))
        target_workspace = _safe(row, header_map.get("target_workspace"))
        target_node_id = _safe_int(_safe(row, header_map.get("target_node_id")))
        relation_type = _safe(row, header_map.get("relation_type")).lower()
        enabled = parse_enabled(_safe(row, header_map.get("enabled")))
        if not any([source_workspace, target_workspace, target_node_id, relation_type, enabled]):
            continue
        rows.append(
            RelatedWorkspaceRow(
                row_index=excel_row_number,
                source_workspace=source_workspace,
                target_workspace=target_workspace,
                target_node_id=target_node_id,
                relation_type=relation_type,
                enabled=enabled,
            )
        )
    return rows


def _safe(row: tuple[Any, ...], col: int | None) -> str:
    if col is None or col < 0 or col >= len(row):
        return ""
    return clean_cell_value(row[col])


def _safe_header(value: Any) -> str:
    return clean_cell_value(value).strip().lower()


def _safe_int(value: str) -> int | None:
    if not value:
        return None
    try:
        return int(float(value))
    except ValueError:
        return None


def parse_enabled(value: str) -> bool:
    return clean_cell_value(value).strip().lower() in ENABLED_TRUE_VALUES


def _resolve_local_path(src: str, local_root: str) -> str:
    if not src:
        return ""
    if local_root and not os.path.isabs(src):
        return str(Path(local_root) / Path(src).name)
    if local_root and os.path.isabs(src):
        # Keep legacy behaviour: a configured local root means the Excel src may
        # contain exported paths, but import should read by basename from root.
        return str(Path(local_root) / Path(src).name)
    return src


def _apply_value_map(value: str, value_map: dict[str, str]) -> str:
    if value in value_map:
        return value_map[value]

    trimmed_value = _normalize_category_value_text(value, casefold=False)
    for source, target in value_map.items():
        if _normalize_category_value_text(source, casefold=False) == trimmed_value:
            return target

    normalized_value = normalize_category_value(value)
    for source, target in value_map.items():
        if normalize_category_value(source) == normalized_value:
            return target
    return value


def normalize_category_value(value: str) -> str:
    return _normalize_category_value_text(value, casefold=True)


def _normalize_category_value_text(value: str, *, casefold: bool) -> str:
    normalized = unicodedata.normalize("NFC", str(value).replace("\u00a0", " "))
    normalized = _WHITESPACE_PATTERN.sub(" ", normalized).strip()
    return normalized.casefold() if casefold else normalized
