from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from .import_engine import ImportExecutionReport
from .models import Issue, PreflightReport


class ReportGenerator:
    def __init__(self, output_dir: str | Path = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write_all(self, report: PreflightReport, *, stem: str = "migration_preflight") -> dict[str, Path]:
        paths = {
            "json": self.write_json(report, f"{stem}.json"),
            "markdown": self.write_markdown(report, f"{stem}.md"),
            "csv": self.write_csv(report.issues, f"{stem}_issues.csv"),
            "xlsx": self.write_xlsx(report.issues, f"{stem}_issues.xlsx"),
        }
        return paths

    def write_execution_all(
        self,
        report: ImportExecutionReport,
        *,
        stem: str = "migration",
        timestamp: str | None = None,
    ) -> dict[str, Path]:
        timestamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_stem = f"{stem}_execution_{timestamp}"
        return {
            "json": self.write_execution_json(report, f"{filename_stem}.json"),
            "markdown": self.write_execution_markdown(report, f"{filename_stem}.md"),
            "csv": self.write_execution_csv(report, f"{filename_stem}.csv"),
            "xlsx": self.write_execution_xlsx(report, f"{filename_stem}.xlsx"),
        }

    def write_json(self, report: PreflightReport, filename: str) -> Path:
        path = self.output_dir / filename
        path.write_text(json.dumps(report.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    def write_markdown(self, report: PreflightReport, filename: str) -> Path:
        path = self.output_dir / filename
        summary = report.summary
        lines = [
            f"# {report.project_name} – Migration Preflight Report",
            "",
            "## Executive summary",
            "",
            f"- **Decision:** {summary.decision}",
            f"- **Readiness score:** {summary.readiness_score}/100",
            f"- **Workspace rows:** {summary.workspace_rows}",
            f"- **File rows:** {summary.file_rows}",
            f"- **Blocking issues:** {summary.blocking_count}",
            f"- **Errors:** {summary.error_count}",
            f"- **Warnings:** {summary.warning_count}",
            f"- **Generated at:** {report.generated_at}",
            f"- **Mode:** {report.mode}",
            "",
            "## Recommendation",
            "",
            self._recommendation(summary.decision),
            "",
            "## Issue overview",
            "",
        ]
        if not report.issues:
            lines.append("No issues found. The input is ready for the next reviewed dry-run step.")
        else:
            lines.extend([
                "| Severity | Code | Row type | Row | Field | Blocking | Message | Suggestion |",
                "|---|---|---|---:|---|---|---|---|",
            ])
            for issue in sorted(report.issues, key=_issue_sort_key):
                lines.append(
                    "| {severity} | `{code}` | {row_type} | {row} | {field} | {blocking} | {message} | {suggestion} |".format(
                        severity=issue.severity.upper(),
                        code=issue.code,
                        row_type=issue.row_type,
                        row=issue.row_index or "",
                        field=issue.field or "",
                        blocking="yes" if issue.blocking else "no",
                        message=_escape_table(issue.message),
                        suggestion=_escape_table(issue.suggestion),
                    )
                )
        lines.extend([
            "",
            "## Next steps",
            "",
            "1. Fix all blocking errors in the Excel or local source folder.",
            "2. Re-run `analyze` until there are no local blocking issues.",
            "3. Run `preflight` with Content Server read-only checks enabled.",
            "4. Only after reviewed approval, run dry-run/import execution in a controlled batch.",
            "",
            "## Metadata",
            "",
            "```json",
            json.dumps(report.metadata, ensure_ascii=False, indent=2),
            "```",
            "",
        ])
        path.write_text("\n".join(lines), encoding="utf-8")
        return path

    def write_csv(self, issues: list[Issue], filename: str) -> Path:
        path = self.output_dir / filename
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=_issue_fieldnames())
            writer.writeheader()
            for issue in sorted(issues, key=_issue_sort_key):
                writer.writerow(issue.to_dict())
        return path

    def write_xlsx(self, issues: list[Issue], filename: str) -> Path:
        path = self.output_dir / filename
        wb = Workbook()
        ws = wb.active
        ws.title = "Issues"
        headers = _issue_fieldnames()
        ws.append(headers)
        for issue in sorted(issues, key=_issue_sort_key):
            data = issue.to_dict()
            ws.append([data.get(header) for header in headers])
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 70)
        wb.save(path)
        return path

    def write_execution_json(self, report: ImportExecutionReport, filename: str) -> Path:
        path = self.output_dir / filename
        path.write_text(json.dumps(report.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    def write_execution_markdown(self, report: ImportExecutionReport, filename: str) -> Path:
        path = self.output_dir / filename
        summary = report.to_dict()["summary"]
        lines = [
            f"# {report.project_name} - Migration Execution Report",
            "",
            "## Summary",
            "",
            f"- **Mode:** {report.mode}",
            f"- **Generated at:** {report.generated_at}",
            f"- **Workbook:** {report.xlsx_path}",
            f"- **Workspaces created:** {summary['ws_created']}",
            f"- **Workspaces existing:** {summary['ws_existing']}",
            f"- **Workspaces failed:** {summary['ws_failed']}",
            f"- **Files uploaded:** {summary['f_uploaded']}",
            f"- **Files versioned:** {summary['f_versioned']}",
            f"- **Files skipped:** {summary['f_skipped']}",
            f"- **Files failed:** {summary['f_failed']}",
            f"- **Related workspaces created:** {summary['related_created']}",
            f"- **Related workspaces existing:** {summary['related_existing']}",
            f"- **Related workspaces skipped:** {summary['related_skipped']}",
            f"- **Related workspaces failed:** {summary['related_failed']}",
            "",
            "## Workspaces",
            "",
            "| Row | Placeholder | CS workspace name | Node id | Status | Error |",
            "|---:|---|---|---:|---|---|",
        ]
        for row in report.stats.workspace_rows:
            lines.append(
                "| {row} | {title} | {name} | {node_id} | {status} | {error} |".format(
                    row=row.row_index,
                    title=_escape_table(row.excel_title),
                    name=_escape_table(row.workspace_name),
                    node_id=row.node_id or "",
                    status=row.status,
                    error=_escape_table(row.error),
                )
            )
        lines.extend([
            "",
            "## Related Workspaces",
            "",
            "| Row | Source | Source node id | Source resolved name | Target | Target node id | Target resolved name | Relation type | Status | Error |",
            "|---:|---|---:|---|---|---:|---|---|---|---|",
        ])
        for row in report.stats.related_rows:
            lines.append(
                "| {row} | {source} | {source_id} | {source_name} | {target} | {target_id} | {target_name} | {rel_type} | {status} | {error} |".format(
                    row=row.row_index,
                    source=_escape_table(row.source_workspace),
                    source_id=row.source_node_id or "",
                    source_name=_escape_table(row.source_resolved_name),
                    target=_escape_table(row.target_workspace),
                    target_id=row.target_node_id or "",
                    target_name=_escape_table(row.target_resolved_name),
                    rel_type=_escape_table(row.relation_type),
                    status=row.status,
                    error=_escape_table(row.error_message),
                )
            )
        lines.extend([
            "",
            "## Files",
            "",
            "| Row | Title | Source | Original target | Remapped target | Parent node id | Status | Error |",
            "|---:|---|---|---|---|---:|---|---|",
        ])
        for row in report.stats.file_rows:
            lines.append(
                "| {row} | {title} | {source} | {original} | {remapped} | {parent} | {status} | {error} |".format(
                    row=row.row_index,
                    title=_escape_table(row.title),
                    source=_escape_table(row.source_path),
                    original=_escape_table(row.original_target_location),
                    remapped=_escape_table(row.remapped_target_location),
                    parent=row.parent_node_id or "",
                    status=row.status,
                    error=_escape_table(row.error),
                )
            )
        if report.stats.details:
            lines.extend(["", "## Details", ""])
            lines.extend(f"- {_escape_table(detail)}" for detail in report.stats.details)
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return path

    def write_execution_csv(self, report: ImportExecutionReport, filename: str) -> Path:
        path = self.output_dir / filename
        fieldnames = _execution_csv_fieldnames()
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in _execution_csv_rows(report):
                writer.writerow(row)
        return path

    def write_execution_xlsx(self, report: ImportExecutionReport, filename: str) -> Path:
        path = self.output_dir / filename
        wb = Workbook()
        ws = wb.active
        ws.title = "Workspaces"
        workspace_headers = ["row_index", "excel_title", "workspace_name", "node_id", "status", "error"]
        ws.append(workspace_headers)
        for row in report.stats.workspace_rows:
            data = row.to_dict()
            ws.append([data.get(header) for header in workspace_headers])
        _format_sheet(ws)

        files_ws = wb.create_sheet("Files")
        file_headers = [
            "row_index",
            "title",
            "source_path",
            "original_target_location",
            "remapped_target_location",
            "parent_node_id",
            "status",
            "error",
        ]
        files_ws.append(file_headers)
        for row in report.stats.file_rows:
            data = row.to_dict()
            files_ws.append([data.get(header) for header in file_headers])
        _format_sheet(files_ws)

        related_ws = wb.create_sheet("RelatedWorkspaces")
        related_headers = [
            "row_index",
            "source_workspace",
            "source_node_id",
            "source_resolved_name",
            "target_workspace",
            "target_node_id",
            "target_resolved_name",
            "relation_type",
            "status",
            "error_message",
        ]
        related_ws.append(related_headers)
        for row in report.stats.related_rows:
            data = row.to_dict()
            related_ws.append([data.get(header) for header in related_headers])
        _format_sheet(related_ws)
        wb.save(path)
        return path

    @staticmethod
    def _recommendation(decision: str) -> str:
        if decision == "GO":
            return "**GO.** The source is suitable for the next dry-run/import planning step."
        if decision == "GO_WITH_WARNINGS":
            return "**GO WITH WARNINGS.** The batch can proceed only after the warnings are reviewed and accepted."
        return "**NO-GO.** Blocking issues must be fixed before any write operation is allowed."


def _issue_fieldnames() -> list[str]:
    return ["severity", "code", "row_type", "row_index", "field", "blocking", "message", "suggestion", "value"]


def _issue_sort_key(issue: Issue) -> tuple[int, int, str]:
    severity_rank = {"error": 0, "warning": 1, "info": 2}
    return (severity_rank.get(issue.severity, 9), issue.row_index or 0, issue.code)


def _escape_table(value: str) -> str:
    return (value or "").replace("|", "\\|").replace("\n", " ")


def _execution_csv_fieldnames() -> list[str]:
    return [
        "record_type",
        "row_index",
        "title",
        "workspace_name",
        "source_path",
        "original_target_location",
        "remapped_target_location",
        "node_id",
        "parent_node_id",
        "source_workspace",
        "source_node_id",
        "source_resolved_name",
        "target_workspace",
        "target_node_id",
        "target_resolved_name",
        "relation_type",
        "status",
        "error",
    ]


def _execution_csv_rows(report: ImportExecutionReport) -> list[dict]:
    rows = []
    for workspace in report.stats.workspace_rows:
        rows.append({
            "record_type": "workspace",
            "row_index": workspace.row_index,
            "title": workspace.excel_title,
            "workspace_name": workspace.workspace_name,
            "source_path": "",
            "original_target_location": "",
            "remapped_target_location": "",
            "node_id": workspace.node_id,
            "parent_node_id": "",
            "source_workspace": "",
            "source_node_id": "",
            "source_resolved_name": "",
            "target_workspace": "",
            "target_node_id": "",
            "target_resolved_name": "",
            "relation_type": "",
            "status": workspace.status,
            "error": workspace.error,
        })
    for related in report.stats.related_rows:
        rows.append({
            "record_type": "related_workspace",
            "row_index": related.row_index,
            "title": "",
            "workspace_name": "",
            "source_path": "",
            "original_target_location": "",
            "remapped_target_location": "",
            "node_id": "",
            "parent_node_id": "",
            "source_workspace": related.source_workspace,
            "source_node_id": related.source_node_id,
            "source_resolved_name": related.source_resolved_name,
            "target_workspace": related.target_workspace,
            "target_node_id": related.target_node_id,
            "target_resolved_name": related.target_resolved_name,
            "relation_type": related.relation_type,
            "status": related.status,
            "error": related.error_message,
        })
    for file in report.stats.file_rows:
        rows.append({
            "record_type": "file",
            "row_index": file.row_index,
            "title": file.title,
            "workspace_name": "",
            "source_path": file.source_path,
            "original_target_location": file.original_target_location,
            "remapped_target_location": file.remapped_target_location,
            "node_id": "",
            "parent_node_id": file.parent_node_id,
            "source_workspace": "",
            "source_node_id": "",
            "source_resolved_name": "",
            "target_workspace": "",
            "target_node_id": "",
            "target_resolved_name": "",
            "relation_type": "",
            "status": file.status,
            "error": file.error,
        })
    return rows


def _format_sheet(ws) -> None:
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 70)
